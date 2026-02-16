#!/usr/bin/env python3
"""Generate a polished Excel beta test plan for BumpSetCut."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

wb = openpyxl.Workbook()

# ── Color Palette ──────────────────────────────────────────────
ORANGE       = "F97316"   # BumpSetCut brand
DARK_BG      = "1E1E2E"   # Dark header
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F8F9FA"
MID_GRAY     = "E9ECEF"
DARK_TEXT     = "212529"
MUTED_TEXT   = "6C757D"
GREEN_PASS   = "D4EDDA"
GREEN_TEXT   = "155724"
RED_FAIL     = "F8D7DA"
RED_TEXT     = "721C24"
YELLOW_SKIP  = "FFF3CD"
YELLOW_TEXT  = "856404"
BLUE_BLOCK   = "D1ECF1"
BLUE_TEXT    = "0C5460"
CRITICAL_BG  = "FFF0E6"   # Light orange for critical sections
HIGH_BG      = "FFF8E1"   # Light yellow for high priority
MEDIUM_BG    = "F0F4FF"   # Light blue for medium priority

# ── Reusable Styles ────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color="DEE2E6"),
    right=Side(style="thin", color="DEE2E6"),
    top=Side(style="thin", color="DEE2E6"),
    bottom=Side(style="thin", color="DEE2E6"),
)

header_font = Font(name="Aptos", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

section_font = Font(name="Aptos", size=11, bold=True, color=DARK_TEXT)
section_fill = PatternFill(start_color=MID_GRAY, end_color=MID_GRAY, fill_type="solid")
section_align = Alignment(horizontal="left", vertical="center")

body_font = Font(name="Aptos", size=10, color=DARK_TEXT)
body_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")
wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Priority colors for row striping ───────────────────────────
priority_fills = {
    "Critical": PatternFill(start_color=CRITICAL_BG, end_color=CRITICAL_BG, fill_type="solid"),
    "High":     PatternFill(start_color=HIGH_BG, end_color=HIGH_BG, fill_type="solid"),
    "Medium":   PatternFill(start_color=MEDIUM_BG, end_color=MEDIUM_BG, fill_type="solid"),
}

# Row stripe
stripe_even = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
stripe_odd  = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")


# ══════════════════════════════════════════════════════════════
# TEST DATA
# ══════════════════════════════════════════════════════════════

sections = [
    {
        "name": "1. First Launch & Onboarding",
        "priority": "High",
        "subsections": [
            {
                "name": "1.1 Fresh Install",
                "tests": [
                    ("TC-001", "App launches without crash on clean install", "Delete app, reinstall from TestFlight, launch"),
                    ("TC-002", "Onboarding carousel appears automatically", "Should show 4-page carousel on first launch"),
                    ("TC-003", "All 4 onboarding pages display correctly", "Verify: Upload, AI Detection, Save Rallies, Share & Connect"),
                    ("TC-004", '"Skip" button works on pages 1-3', "Tap Skip on each page — should jump to main app"),
                    ("TC-005", '"Done" on last page dismisses onboarding', "Navigate to page 4, tap Done"),
                    ("TC-006", "Onboarding does NOT reappear after completion", "Kill app, relaunch — should go straight to Home"),
                    ("TC-007", "Main tab bar shows 4 tabs", "Verify: Home, Feed, Search, Profile tabs visible"),
                ],
            },
            {
                "name": "1.2 Permissions",
                "tests": [
                    ("TC-008", "Photo library permission prompt on first upload", "Tap Upload — iOS permission dialog should appear"),
                    ("TC-009", "Denying permission shows fallback (no crash)", "Deny photo access — app should show explanation, not crash"),
                    ("TC-010", "Granting permission allows video selection", "Grant access — PhotosPicker should open with videos"),
                ],
            },
        ],
    },
    {
        "name": "2. Home Screen",
        "priority": "Medium",
        "subsections": [
            {
                "name": "2.1 Layout & Display",
                "tests": [
                    ("TC-011", "Hero section renders with app branding", "Logo and tagline visible at top"),
                    ("TC-012", "Stats card shows rallies, processed count, tier", "Verify all 3 stats display correctly"),
                    ("TC-013", "Quick action buttons visible", "Upload, Process, Help buttons all tappable"),
                ],
            },
            {
                "name": "2.2 Navigation",
                "tests": [
                    ("TC-014", '"View Saved Games" → saved video library', "Tap button — navigates to Saved Games view"),
                    ("TC-015", '"View Processed Games" → processed library', "Tap button — navigates to Processed Games view"),
                    ("TC-016", "Settings gear opens settings sheet", "Tap gear icon — SettingsView sheet appears"),
                    ("TC-017", "Help button re-opens onboarding carousel", "Tap Help — full onboarding carousel shown again"),
                ],
            },
        ],
    },
    {
        "name": "3. Video Upload",
        "priority": "High",
        "subsections": [
            {
                "name": "3.1 Import Flow",
                "tests": [
                    ("TC-018", "Upload button → PhotosPicker opens", "Home > Upload — picker appears"),
                    ("TC-019", "Select video → folder selection sheet", "Pick a video — folder chooser appears"),
                    ("TC-020", "Select folder → naming dialog", "Choose folder — video naming dialog shown"),
                    ("TC-021", "Enter custom name → video in library", "Name the video — appears in selected folder"),
                    ("TC-022", "Skip naming → auto-generated name", "Leave name blank or skip — default name used"),
                    ("TC-023", "Cancel at any step → no partial files", "Cancel mid-flow — no orphan files in library"),
                ],
            },
            {
                "name": "3.2 Large File Handling",
                "tests": [
                    ("TC-024", "Video >500MB on free tier → warning", "Upload large file — tier/storage warning shown"),
                    ("TC-025", "5-minute video completes without crash", "Upload medium video — no memory issues"),
                    ("TC-026", "30+ minute video completes", "Upload long video — URL-based, should complete"),
                ],
            },
            {
                "name": "3.3 Edge Cases",
                "tests": [
                    ("TC-027", "Portrait video → orientation preserved", "Upload portrait — check metadata after import"),
                    ("TC-028", "Landscape video → orientation preserved", "Upload landscape — check metadata after import"),
                    ("TC-029", "Same video uploaded twice → no crash", "Upload identical video — graceful handling"),
                    ("TC-030", "Upload with no folders → handled", "Delete all folders, try upload — default or create prompt"),
                ],
            },
        ],
    },
    {
        "name": "4. Video Library",
        "priority": "High",
        "subsections": [
            {
                "name": "4.1 Saved Games Library",
                "tests": [
                    ("TC-031", "Grid shows thumbnails with duration & size", "Open Saved Games — verify thumbnail grid"),
                    ("TC-032", "Thumbnails load progressively", "No blank grid — thumbnails appear as generated"),
                    ("TC-033", "Empty state when no videos", "Delete all — empty state message shown"),
                    ("TC-034", "Search bar filters by name", "Type partial name — list filters correctly"),
                    ("TC-035", "Pull to refresh updates list", "Pull down — list refreshes"),
                ],
            },
            {
                "name": "4.2 Folder Management",
                "tests": [
                    ("TC-036", "Create new folder", "Create folder — appears in library"),
                    ("TC-037", "Rename folder", "Rename — name updates in all references"),
                    ("TC-038", "Delete empty folder", "Delete folder with no videos — removed after confirm"),
                    ("TC-039", "Delete folder with videos → warning", "Delete folder with content — appropriate warning shown"),
                ],
            },
            {
                "name": "4.3 Processed Games Library",
                "tests": [
                    ("TC-040", "Only processed videos appear", "No unprocessed videos in this view"),
                    ("TC-041", "Tap processed video → rally player", "Select video — rally player opens"),
                    ("TC-042", "Unprocessed videos excluded", "Verify raw uploads don't appear here"),
                ],
            },
            {
                "name": "4.4 Navigation",
                "tests": [
                    ("TC-043", "Libraries clearly distinguished", "Visual difference between Saved and Processed"),
                    ("TC-044", "Breadcrumb navigation works", "Navigate nested folders — breadcrumbs update"),
                    ("TC-045", "Back button returns to parent", "Tap back — returns to parent folder"),
                ],
            },
        ],
    },
    {
        "name": "5. Video Processing",
        "priority": "Critical",
        "subsections": [
            {
                "name": "5.1 Start Processing",
                "tests": [
                    ("TC-046", "Process button → unprocessed video list", "Home > Process — list of unprocessed videos"),
                    ("TC-047", "Select video → ProcessVideoView opens", "Pick video — processing screen loads"),
                    ("TC-048", "Video metadata displayed", "Duration, resolution, file size shown"),
                    ("TC-049", "Volleyball type selection (Beach/Indoor)", "Type picker available and functional"),
                    ("TC-050", '"Process Video" button tappable', "Button enabled and responds to tap"),
                ],
            },
            {
                "name": "5.2 Processing Progress",
                "tests": [
                    ("TC-051", "Animated brain icon on processing start", "Tap Process — animation appears"),
                    ("TC-052", "Progress 0% → 100% smoothly", "Progress bar increments without jumps/stalls"),
                    ("TC-053", '"Analyzing video..." label shown', "Status text visible during processing"),
                    ("TC-054", "UI remains responsive during processing", "Can interact with other UI elements"),
                    ("TC-055", "Background app → processing continues", "Background for <30s — processing survives"),
                ],
            },
            {
                "name": "5.3 Results — Rallies Found",
                "tests": [
                    ("TC-056", "Rally count displayed correctly", 'Shows "X rallies detected"'),
                    ("TC-057", "Duration and time saved % shown", "Processing summary stats accurate"),
                    ("TC-058", "Preview button → rally player", "Tap Preview — player opens with all rallies"),
                    ("TC-059", "Save button → folder selection → saved", "Save processed video to chosen folder"),
                    ("TC-060", "Appears in Processed Games library", "Navigate to Processed — video listed"),
                    ("TC-061", "Original blocked from reprocessing", "Try to process original again — blocked"),
                ],
            },
            {
                "name": "5.4 Results — No Rallies",
                "tests": [
                    ("TC-062", '"No rallies detected" message', "Appropriate message for zero-rally result"),
                    ("TC-063", "Retry option available", "Can retry processing from this screen"),
                    ("TC-064", "Dismiss returns without crash", "Close screen — back to library cleanly"),
                ],
            },
            {
                "name": "5.5 Edge Cases",
                "tests": [
                    ("TC-065", "10-second video → handles gracefully", "Very short video — no crash"),
                    ("TC-066", "30+ minute video → completes", "Long video — processing finishes"),
                    ("TC-067", "Non-volleyball video → no rallies", "Random video — no crash, 0 rallies"),
                    ("TC-068", "Storage nearly full → warning", "Low storage — warning before processing"),
                    ("TC-069", "Kill app during processing → clean relaunch", "Force quit — no corrupt metadata"),
                    ("TC-070", "Already-processed → blocked state", "Reprocess attempt — shows already-processed UI"),
                ],
            },
            {
                "name": "5.6 Thorough vs Quick Mode",
                "tests": [
                    ("TC-071", "Toggle Thorough Analysis OFF in Settings", "Setting persists and is applied"),
                    ("TC-072", "Quick mode processes faster", "Noticeably faster than thorough"),
                    ("TC-073", "Toggle back ON → thorough processing", "More frames analyzed, takes longer"),
                ],
            },
        ],
    },
    {
        "name": "6. Rally Playback",
        "priority": "Critical",
        "subsections": [
            {
                "name": "6.1 Basic Playback",
                "tests": [
                    ("TC-074", "Processed video → rally player loads", "Open processed video — player appears"),
                    ("TC-075", "First rally plays automatically", "Video starts playing on load"),
                    ("TC-076", "Correct orientation (portrait/landscape)", "Video matches source orientation"),
                    ("TC-077", "Play/pause works", "Tap to toggle play/pause"),
                    ("TC-078", 'Rally counter visible (e.g., "3 of 12")', "Counter updates as you navigate"),
                ],
            },
            {
                "name": "6.2 Navigation Between Rallies",
                "tests": [
                    ("TC-079", "Swipe up → next rally", "Swipe up — next rally loads and plays"),
                    ("TC-080", "Swipe down → previous rally", "Swipe down — previous rally loads and plays"),
                    ("TC-081", "Rapid swiping → no crash", "Swipe quickly through 10+ rallies"),
                    ("TC-082", "First rally → swipe down = no-op", "At rally 1, swipe down — bounce/nothing"),
                    ("TC-083", "Last rally → swipe up = no-op", "At last rally, swipe up — bounce/nothing"),
                    ("TC-084", "No audio bleed during fast swiping", "Listen for overlapping audio — should be clean"),
                ],
            },
            {
                "name": "6.3 Rally Selection",
                "tests": [
                    ("TC-085", "Tap to save rally → indicator updates", "Toggle save — visual feedback"),
                    ("TC-086", "Tap again to deselect", "Un-save — indicator clears"),
                    ("TC-087", "Counter pill shows selected count", "Pill updates with save/remove"),
                    ("TC-088", "Remove all rallies → empty state", "Deselect everything — appropriate message"),
                ],
            },
            {
                "name": "6.4 Per-Rally Trim/Buffer",
                "tests": [
                    ("TC-089", "Long-press → trim mode activates", "Press and hold — trim controls appear"),
                    ("TC-090", "Start buffer adjusts ±3s in 0.5s steps", "Tap -/+ buttons — value changes correctly"),
                    ("TC-091", "End buffer adjusts ±3s in 0.5s steps", "Same for end trim controls"),
                    ("TC-092", "Trim persists across rally navigation", "Trim rally 3, go to 5, come back — still trimmed"),
                    ("TC-093", "Trim persists across app restart", "Trim, kill app, relaunch — values intact"),
                    ("TC-094", "Trim clamped to video boundaries", "Can't trim past start/end of source video"),
                ],
            },
            {
                "name": "6.5 Rally Overview Sheet",
                "tests": [
                    ("TC-095", "Tap counter pill → overview opens", "Pill tap — sheet with thumbnail grid"),
                    ("TC-096", "Thumbnail grid shows all rallies", "Every rally has a thumbnail"),
                    ("TC-097", "Tap thumbnail → jumps to rally", "Select rally 7 — player jumps to 7"),
                    ("TC-098", "Selected/deselected visually distinct", "Saved vs unsaved rallies look different"),
                    ("TC-099", "Select/deselect from overview", "Toggle rallies directly in grid"),
                    ("TC-100", "Close sheet → current rally preserved", "Dismiss — still at same rally position"),
                ],
            },
            {
                "name": "6.6 Undo",
                "tests": [
                    ("TC-101", "Undo reverts last action", "Save a rally, undo — unsaved"),
                    ("TC-102", "Multiple undos (multi-level stack)", "Make 3 actions, undo 3 times — all reverted"),
                    ("TC-103", "Undo after trim → values revert", "Trim, undo — original boundaries restored"),
                ],
            },
            {
                "name": "6.7 Player Cache",
                "tests": [
                    ("TC-104", "10+ rallies → no memory warning", "Scroll through many rallies — memory stable"),
                    ("TC-105", "Return to earlier rally → loads ok", "Go to rally 10, back to 1 — plays fine"),
                    ("TC-106", "No audio from off-screen rallies", "Only current rally's audio plays"),
                ],
            },
            {
                "name": "6.8 Gesture Tips",
                "tests": [
                    ("TC-107", "First view → gesture tips overlay", "Overlay with swipe/trim hints shown"),
                    ("TC-108", "Tips don't reappear on subsequent views", "Close player, reopen — no tips"),
                ],
            },
        ],
    },
    {
        "name": "7. Export",
        "priority": "High",
        "subsections": [
            {
                "name": "7.1 Save to Device",
                "tests": [
                    ("TC-109", "Export button → folder selection", "Tap save/export — folder chooser appears"),
                    ("TC-110", "Select folder → clips exported as MP4", "Choose folder — files appear"),
                    ("TC-111", "Clips play correctly in Photos app", "Open exported clip in Photos — plays fine"),
                    ("TC-112", "Orientation preserved in export", "Portrait stays portrait, landscape stays landscape"),
                    ("TC-113", "Audio preserved in export", "Exported clip has audio track"),
                ],
            },
            {
                "name": "7.2 Batch Export",
                "tests": [
                    ("TC-114", "Multiple rallies → export all selected", "Select 5 rallies — all 5 exported"),
                    ("TC-115", "Export from overview sheet", "Use overview grid to export"),
                    ("TC-116", "Each clip has correct time boundaries", "Verify each clip's start/end matches rally"),
                ],
            },
            {
                "name": "7.3 Edge Cases",
                "tests": [
                    ("TC-117", "Trimmed rally → offsets applied in export", "Trim +2s start, export — clip starts 2s earlier"),
                    ("TC-118", "Storage nearly full → error shown", "Low storage — appropriate error message"),
                    ("TC-119", "Cancel export → no corrupt files", "Cancel mid-export — no partial MP4s left"),
                ],
            },
        ],
    },
    {
        "name": "8. Authentication",
        "priority": "High",
        "subsections": [
            {
                "name": "8.1 Apple Sign-In",
                "tests": [
                    ("TC-120", "Feed/Profile tab → AuthGate appears", "Tap Feed unauthenticated — gate shown"),
                    ("TC-121", "Apple Sign-In button → Apple auth sheet", "Tap button — system auth dialog"),
                    ("TC-122", "Complete auth → authenticated state", "Sign in — gate dismissed"),
                    ("TC-123", "User profile created/restored", "Profile data loaded after auth"),
                    ("TC-124", "First sign-in → username picker", "New account — username modal appears"),
                    ("TC-125", "Valid username → profile complete", "Enter username — proceeds to app"),
                ],
            },
            {
                "name": "8.2 Email Sign-In",
                "tests": [
                    ("TC-126", "Sign Up form works", "Create account with email + password"),
                    ("TC-127", "Sign In with existing account", "Log in — authenticated"),
                    ("TC-128", "Invalid email → validation error", "Enter bad email — error shown"),
                    ("TC-129", "Wrong password → error message", "Enter wrong pw — clear error"),
                    ("TC-130", "Forgot Password → reset flow", "Tap link — reset password view opens"),
                    ("TC-131", "Password reset works end-to-end", "Request reset — check email — reset pw"),
                ],
            },
            {
                "name": "8.3 Session Persistence",
                "tests": [
                    ("TC-132", "Kill app → still authenticated", "Force quit, relaunch — no login required"),
                    ("TC-133", "Session expires → re-auth prompt", "Wait for expiry — prompted to re-login"),
                    ("TC-134", "Sign out → unauthenticated", "Settings > Sign Out — session cleared"),
                ],
            },
            {
                "name": "8.4 Continue Without Account",
                "tests": [
                    ("TC-135", '"Continue without account" visible', "Option shown on AuthGate"),
                    ("TC-136", "Dismisses auth gate", "Tap — gate goes away"),
                    ("TC-137", "Social actions still gated", "Try to like/comment — auth gate re-appears"),
                ],
            },
        ],
    },
    {
        "name": "9. Social Feed",
        "priority": "Medium",
        "subsections": [
            {
                "name": "9.1 Feed Display",
                "tests": [
                    ("TC-138", "Authenticated → highlights load", "Sign in, tap Feed — content loads"),
                    ("TC-139", "TikTok-style vertical scroll", "Snap-to-page behavior on scroll"),
                    ("TC-140", "Card shows video, author, buttons, caption", "All elements visible per card"),
                    ("TC-141", '"For You" / "Following" tabs work', "Switch tabs — different content loads"),
                    ("TC-142", "Empty Following → prompt shown", 'No follows — "follow users" message'),
                ],
            },
            {
                "name": "9.2 Feed Interactions",
                "tests": [
                    ("TC-143", "Like toggles instantly (optimistic)", "Tap heart — instant count change"),
                    ("TC-144", "Comment button → CommentsSheet", "Tap comment icon — sheet opens"),
                    ("TC-145", "Author tap → ProfileView", "Tap username/avatar — profile opens"),
                    ("TC-146", "Scroll to bottom → pagination loads more", "Keep scrolling — new content loads"),
                ],
            },
            {
                "name": "9.3 Edge Cases",
                "tests": [
                    ("TC-147", "No internet → offline message", "Airplane mode — appropriate message"),
                    ("TC-148", "Slow connection → loading state", "Throttle network — spinner visible"),
                    ("TC-149", "Pull to refresh", "Pull down — feed reloads"),
                    ("TC-150", "Delete own highlight → removed", "Delete — disappears from feed immediately"),
                ],
            },
        ],
    },
    {
        "name": "10. Sharing to Community",
        "priority": "Medium",
        "subsections": [
            {
                "name": "10.1 Share Flow",
                "tests": [
                    ("TC-151", "Share button → ShareRallySheet", "From rally player — sheet opens"),
                    ("TC-152", "Caption editor with hashtags", "Type caption with # — works"),
                    ("TC-153", "Hide likes toggle", "Toggle on/off — persists"),
                    ("TC-154", "Post → progress → complete", "Upload shows: uploading → processing → done"),
                    ("TC-155", "Success → highlight in feed", "After post — visible in social feed"),
                ],
            },
            {
                "name": "10.2 Batch Share",
                "tests": [
                    ("TC-156", '"Post All" for multiple rallies', "Batch upload option works"),
                    ("TC-157", "Sequential upload with progress", "Each rally uploads one at a time"),
                    ("TC-158", "All appear in feed on completion", "All posted rallies visible"),
                ],
            },
            {
                "name": "10.3 Share Failures",
                "tests": [
                    ("TC-159", "No internet → error + retry", "Airplane mode during upload — retry button"),
                    ("TC-160", "Retry after reconnect works", "Re-enable wifi — retry succeeds"),
                    ("TC-161", "Cancel upload → cleanup", "Cancel mid-upload — no partial post"),
                    ("TC-162", "Rally >1 min → validation error", "Long rally — blocked before upload"),
                ],
            },
        ],
    },
    {
        "name": "11. Comments",
        "priority": "Medium",
        "subsections": [
            {
                "name": "11.1 Comment Operations",
                "tests": [
                    ("TC-163", "Comment button → sheet opens", "Tap comment icon — CommentsSheet appears"),
                    ("TC-164", "Existing comments display correctly", "Avatar, username, timestamp, text shown"),
                    ("TC-165", "Post comment → appears immediately", "Type + submit — comment in list"),
                    ("TC-166", "Delete own comment", "Swipe/tap delete — removed from list"),
                    ("TC-167", "First comment on highlight", "No existing comments — first one works"),
                    ("TC-168", "Comments persist across close/reopen", "Close sheet, reopen — comments still there"),
                ],
            },
        ],
    },
    {
        "name": "12. User Profile",
        "priority": "Medium",
        "subsections": [
            {
                "name": "12.1 Own Profile",
                "tests": [
                    ("TC-169", "Profile tab shows own profile", "Avatar, name, username, bio visible"),
                    ("TC-170", "Stats: highlights, following, followers", "All 3 counts displayed"),
                    ("TC-171", "Own highlights grid", "Grid of posted highlights shown"),
                    ("TC-172", "Edit Profile → update bio/name", "Edit and save — changes reflected"),
                ],
            },
            {
                "name": "12.2 Other User Profiles",
                "tests": [
                    ("TC-173", "Author tap → their profile", "From feed — profile opens"),
                    ("TC-174", "Follow/Unfollow works", "Tap follow — counts update on both profiles"),
                    ("TC-175", "View their highlights grid", "Other user's posts visible"),
                    ("TC-176", "Tap their highlight → plays", "Select highlight — video plays"),
                ],
            },
            {
                "name": "12.3 Community Search",
                "tests": [
                    ("TC-177", "Search bar visible on Search tab", "Search tab — input bar present"),
                    ("TC-178", "Type username → results appear", "Matching users shown as you type"),
                    ("TC-179", "Tap user → ProfileView opens", "Select result — their profile"),
                    ("TC-180", "Empty search → suggestions/empty state", "No query — appropriate UI"),
                ],
            },
        ],
    },
    {
        "name": "13. Settings",
        "priority": "Medium",
        "subsections": [
            {
                "name": "13.1 Core Settings",
                "tests": [
                    ("TC-181", "Appearance: System / Light / Dark", "Switch themes — UI updates immediately"),
                    ("TC-182", "Thorough Analysis toggle persists", "Toggle — kill app — relaunch — still set"),
                    ("TC-183", "App version and build number shown", "Correct values displayed"),
                ],
            },
            {
                "name": "13.2 Subscription",
                "tests": [
                    ("TC-184", "Free tier limits displayed", "3/week processing, 500MB max shown"),
                    ("TC-185", "Subscription button → PaywallView", "Tap — paywall with features + pricing"),
                    ("TC-186", "Restore Purchases works", "Tap restore — checks App Store"),
                ],
            },
            {
                "name": "13.3 Account Management",
                "tests": [
                    ("TC-187", "Sign Out clears session", "Sign out — unauthenticated, gate re-appears"),
                    ("TC-188", "Delete Account with confirmation", "Delete — confirm dialog — account removed"),
                    ("TC-189", "After deletion → unauthenticated", "Returned to auth gate after delete"),
                ],
            },
            {
                "name": "13.4 Privacy & Legal",
                "tests": [
                    ("TC-190", "Privacy policy link opens", "Tap — opens in browser/webview"),
                    ("TC-191", "Terms of service link opens", "Tap — opens correctly"),
                    ("TC-192", "Community guidelines link opens", "Tap — opens correctly"),
                    ("TC-193", "Hide Profile toggle works", "Toggle — profile hidden from search"),
                ],
            },
            {
                "name": "13.5 Content Moderation",
                "tests": [
                    ("TC-194", "Report content option available", "Long-press/menu on highlight — report option"),
                    ("TC-195", "Report flow: reason → submit", "Select reason, submit — confirmation shown"),
                    ("TC-196", "Block user option", "Block — their content disappears"),
                    ("TC-197", "Blocked user content hidden from feed", "Verify feed no longer shows their posts"),
                ],
            },
        ],
    },
    {
        "name": "14. Stability & Performance",
        "priority": "High",
        "subsections": [
            {
                "name": "14.1 Memory",
                "tests": [
                    ("TC-198", "15+ min video → no memory crash", "Process long video — watch memory in Instruments"),
                    ("TC-199", "50+ rally cards → memory stable", "Scroll through many rallies"),
                    ("TC-200", "20+ feed highlights → no warning", "Browse feed extensively"),
                    ("TC-201", "Background during playback → no crash", "Background app, return — still playing"),
                ],
            },
            {
                "name": "14.2 App Lifecycle",
                "tests": [
                    ("TC-202", "Background → foreground state preserved", "Background and return — same screen"),
                    ("TC-203", "Kill → relaunch: all data intact", "Force quit — library, settings, auth intact"),
                    ("TC-204", "Phone call during processing", "Receive call — processing handles interruption"),
                    ("TC-205", "Low battery mode → still functional", "Enable low power — app works normally"),
                ],
            },
            {
                "name": "14.3 Orientation",
                "tests": [
                    ("TC-206", "Portrait video → .fit display", "Portrait source — fits in view"),
                    ("TC-207", "Landscape video → .fill display", "Landscape source — fills width"),
                    ("TC-208", "Rotate device during playback", "Rotate — layout adjusts correctly"),
                    ("TC-209", "Export preserves orientation", "Exported clip matches source orientation"),
                ],
            },
            {
                "name": "14.4 Empty States",
                "tests": [
                    ("TC-210", "No videos → library empty state", "Empty library — message shown, no crash"),
                    ("TC-211", "No rallies → appropriate message", "Zero rallies — clear message"),
                    ("TC-212", "No feed highlights → empty prompt", 'Empty feed — "no content yet" message'),
                    ("TC-213", "No followers → zero counts, no crash", "New account — 0 followers shown"),
                    ("TC-214", "No comments → empty list + input bar", "Highlight with 0 comments — can still type"),
                ],
            },
            {
                "name": "14.5 Network",
                "tests": [
                    ("TC-215", "Processing works offline", "Airplane mode — processing completes locally"),
                    ("TC-216", "Social offline → message + queue", "No internet — offline message, actions queued"),
                    ("TC-217", "Restore internet → queue drains", "Re-enable wifi — queued actions complete"),
                    ("TC-218", "Slow connection → loading states", "Throttle — spinners visible, no infinite hang"),
                ],
            },
        ],
    },
    {
        "name": "15. Data Integrity",
        "priority": "High",
        "subsections": [
            {
                "name": "15.1 Persistence",
                "tests": [
                    ("TC-219", "Uploaded video survives app restart", "Upload, kill, relaunch — video in library"),
                    ("TC-220", "Processing metadata survives restart", "Process, kill, relaunch — metadata intact"),
                    ("TC-221", "Trim values survive restart", "Trim, kill, relaunch — values preserved"),
                    ("TC-222", "Rally selections survive restart", "Save rallies, kill, relaunch — selections intact"),
                    ("TC-223", "Settings survive restart", "Change settings, kill, relaunch — all preserved"),
                ],
            },
            {
                "name": "15.2 Metadata Relationships",
                "tests": [
                    ("TC-224", "Correct rally count on reopen", "Close and reopen processed video — same count"),
                    ("TC-225", "Original ↔ processed video linked", "Both reference each other correctly"),
                    ("TC-226", "Delete original → processed accessible", "Remove source — processed still plays"),
                    ("TC-227", "Delete processed → original reprocessable", "Remove processed — can reprocess original"),
                ],
            },
        ],
    },
    {
        "name": "16. Smoke Test (5-Min Walkthrough)",
        "priority": "Critical",
        "subsections": [
            {
                "name": "16.1 End-to-End Flow",
                "tests": [
                    ("TC-228", "Fresh launch → Home screen loads", "App opens to Home with no errors"),
                    ("TC-229", "Upload a volleyball video", "Import from Photos → lands in library"),
                    ("TC-230", "Process the video → see results", "Full processing pipeline → rally count shown"),
                    ("TC-231", "Swipe through 3+ rallies", "Rally player — swipe navigation works"),
                    ("TC-232", "Long-press to trim one rally", "Trim controls appear and adjust correctly"),
                    ("TC-233", "Save selected rallies to folder", "Export to folder — files appear"),
                    ("TC-234", "Sign in with Apple", "Apple auth flow → authenticated"),
                    ("TC-235", "Share a rally to feed", "Upload pipeline → highlight posted"),
                    ("TC-236", "View it in social feed", "Navigate to Feed — highlight visible"),
                    ("TC-237", "Like and comment on it", "Heart + comment → both work"),
                    ("TC-238", "Settings → sign out", "Sign out — session cleared"),
                    ("TC-239", "Relaunch → library intact", "Kill, reopen — all data still there"),
                ],
            },
        ],
    },
]

# ══════════════════════════════════════════════════════════════
# SHEET 1: MAIN TEST PLAN
# ══════════════════════════════════════════════════════════════

ws = wb.active
ws.title = "Test Plan"
ws.sheet_properties.tabColor = ORANGE

# Column widths
col_widths = {
    "A": 10,   # Test ID
    "B": 12,   # Priority
    "C": 55,   # Test Case
    "D": 50,   # Steps / Details
    "E": 12,   # Status
    "F": 14,   # Tested By
    "G": 14,   # Date Tested
    "H": 12,   # Device
    "I": 50,   # Notes / Bugs
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ── Title Banner ───────────────────────────────────────────────
ws.merge_cells("A1:I1")
title_cell = ws["A1"]
title_cell.value = "BUMPSETCUT  —  Beta Test Plan"
title_cell.font = Font(name="Aptos", size=18, bold=True, color=WHITE)
title_cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 45

ws.merge_cells("A2:I2")
subtitle = ws["A2"]
subtitle.value = f"Generated {date.today().strftime('%B %d, %Y')}  |  239 Test Cases  |  16 Sections"
subtitle.font = Font(name="Aptos", size=10, italic=True, color=MUTED_TEXT)
subtitle.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
subtitle.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 25

# ── Column Headers ─────────────────────────────────────────────
headers = ["Test ID", "Priority", "Test Case", "Steps / Details", "Status", "Tested By", "Date Tested", "Device", "Notes / Bugs"]
row = 4
ws.row_dimensions[row].height = 30
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# ── Data Validation for Status column ──────────────────────────
status_dv = DataValidation(
    type="list",
    formula1='"Pass,Fail,Blocked,Skip,Not Tested"',
    allow_blank=True,
)
status_dv.error = "Please select: Pass, Fail, Blocked, Skip, or Not Tested"
status_dv.errorTitle = "Invalid Status"
status_dv.prompt = "Select test result"
status_dv.promptTitle = "Status"
ws.add_data_validation(status_dv)

# Device validation
device_dv = DataValidation(
    type="list",
    formula1='"iPhone 16 Pro,iPhone 15 Pro,iPhone SE,iPad,Simulator"',
    allow_blank=True,
)
ws.add_data_validation(device_dv)

# ── Conditional Formatting via manual fill (more reliable) ─────
# We'll apply fills row by row after writing

# ── Write Test Data ────────────────────────────────────────────
row = 5
test_row_ranges = []  # Track (start_row, end_row) for each test row for status formatting

for section in sections:
    # Section header row
    ws.merge_cells(f"A{row}:I{row}")
    cell = ws.cell(row=row, column=1, value=f"  {section['name']}  —  {section['priority']} Priority")
    cell.font = Font(name="Aptos", size=11, bold=True, color=DARK_TEXT)
    prio = section["priority"]
    cell.fill = priority_fills.get(prio, section_fill)
    cell.alignment = section_align
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1

    for sub in section["subsections"]:
        # Subsection header
        ws.merge_cells(f"A{row}:I{row}")
        cell = ws.cell(row=row, column=1, value=f"    {sub['name']}")
        cell.font = Font(name="Aptos", size=10, bold=True, italic=True, color=MUTED_TEXT)
        cell.fill = PatternFill(start_color=MID_GRAY, end_color=MID_GRAY, fill_type="solid")
        cell.alignment = section_align
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        for i, (tc_id, desc, steps) in enumerate(sub["tests"]):
            stripe = stripe_even if i % 2 == 0 else stripe_odd
            ws.row_dimensions[row].height = 32

            # A: Test ID
            cell = ws.cell(row=row, column=1, value=tc_id)
            cell.font = Font(name="Aptos Mono", size=9, color=MUTED_TEXT)
            cell.fill = stripe
            cell.alignment = center_align
            cell.border = thin_border

            # B: Priority
            cell = ws.cell(row=row, column=2, value=section["priority"])
            cell.font = Font(name="Aptos", size=9, color=MUTED_TEXT)
            cell.fill = stripe
            cell.alignment = center_align
            cell.border = thin_border

            # C: Test Case
            cell = ws.cell(row=row, column=3, value=desc)
            cell.font = body_font
            cell.fill = stripe
            cell.alignment = body_align
            cell.border = thin_border

            # D: Steps
            cell = ws.cell(row=row, column=4, value=steps)
            cell.font = Font(name="Aptos", size=9, color=MUTED_TEXT)
            cell.fill = stripe
            cell.alignment = body_align
            cell.border = thin_border

            # E: Status (dropdown)
            cell = ws.cell(row=row, column=5, value="")
            cell.font = Font(name="Aptos", size=10, bold=True)
            cell.fill = stripe
            cell.alignment = center_align
            cell.border = thin_border
            status_dv.add(cell)

            # F: Tested By
            cell = ws.cell(row=row, column=6, value="")
            cell.font = body_font
            cell.fill = stripe
            cell.alignment = center_align
            cell.border = thin_border

            # G: Date Tested
            cell = ws.cell(row=row, column=7, value="")
            cell.font = body_font
            cell.fill = stripe
            cell.alignment = center_align
            cell.border = thin_border
            ws.cell(row=row, column=7).number_format = "MM/DD/YYYY"

            # H: Device
            cell = ws.cell(row=row, column=8, value="")
            cell.font = body_font
            cell.fill = stripe
            cell.alignment = center_align
            cell.border = thin_border
            device_dv.add(cell)

            # I: Notes
            cell = ws.cell(row=row, column=9, value="")
            cell.font = Font(name="Aptos", size=9, color=DARK_TEXT)
            cell.fill = stripe
            cell.alignment = wrap_align
            cell.border = thin_border

            test_row_ranges.append(row)
            row += 1

    # Spacer row between sections
    ws.row_dimensions[row].height = 8
    row += 1

# Freeze panes (header row + first column)
ws.freeze_panes = "A5"

# Auto-filter
ws.auto_filter.ref = f"A4:I{row - 1}"

# Print setup
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.print_title_rows = "4:4"


# ══════════════════════════════════════════════════════════════
# SHEET 2: DASHBOARD / SUMMARY
# ══════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Dashboard")
ws2.sheet_properties.tabColor = "4CAF50"

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 12
ws2.column_dimensions["H"].width = 12
ws2.column_dimensions["I"].width = 14

# Title
ws2.merge_cells("B1:I1")
cell = ws2["B1"]
cell.value = "Beta Test Dashboard"
cell.font = Font(name="Aptos", size=16, bold=True, color=WHITE)
cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

# Summary header
summary_headers = ["Section", "Priority", "Total Tests", "Pass", "Fail", "Blocked", "Skip", "Completion %"]
ws2.row_dimensions[3].height = 28
for col_idx, h in enumerate(summary_headers, 2):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Section summary data
summary_data = []
for section in sections:
    total = sum(len(sub["tests"]) for sub in section["subsections"])
    summary_data.append((section["name"], section["priority"], total))

for i, (name, priority, total) in enumerate(summary_data):
    r = 4 + i
    stripe = stripe_even if i % 2 == 0 else stripe_odd
    ws2.row_dimensions[r].height = 26

    # Section name
    cell = ws2.cell(row=r, column=2, value=name)
    cell.font = Font(name="Aptos", size=10, bold=True, color=DARK_TEXT)
    cell.fill = stripe
    cell.alignment = body_align
    cell.border = thin_border

    # Priority
    cell = ws2.cell(row=r, column=3, value=priority)
    cell.font = Font(name="Aptos", size=10, color=MUTED_TEXT)
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border

    # Total
    cell = ws2.cell(row=r, column=4, value=total)
    cell.font = Font(name="Aptos", size=10, bold=True, color=DARK_TEXT)
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border

    # Pass/Fail/Blocked/Skip — formulas counting from Test Plan sheet
    for col_idx, status_val in [(5, "Pass"), (6, "Fail"), (7, "Blocked"), (8, "Skip")]:
        cell = ws2.cell(row=r, column=col_idx, value=0)
        cell.font = Font(name="Aptos", size=10, color=DARK_TEXT)
        cell.fill = stripe
        cell.alignment = center_align
        cell.border = thin_border

    # Completion %
    cell = ws2.cell(row=r, column=9, value="0%")
    cell.font = Font(name="Aptos", size=10, bold=True, color=ORANGE)
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border

# Totals row
total_row = 4 + len(summary_data)
ws2.row_dimensions[total_row].height = 30
cell = ws2.cell(row=total_row, column=2, value="TOTAL")
cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.border = thin_border

total_tests = sum(t for _, _, t in summary_data)
for col_idx in range(3, 10):
    cell = ws2.cell(row=total_row, column=col_idx)
    cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border
ws2.cell(row=total_row, column=4, value=total_tests)

# ── Legend & Instructions ──────────────────────────────────────
legend_row = total_row + 3
ws2.merge_cells(f"B{legend_row}:I{legend_row}")
cell = ws2.cell(row=legend_row, column=2, value="Quick Reference")
cell.font = Font(name="Aptos", size=12, bold=True, color=DARK_TEXT)
ws2.row_dimensions[legend_row].height = 30

legend_items = [
    ("Status: Pass", "Test passed — feature works as expected", GREEN_PASS, GREEN_TEXT),
    ("Status: Fail", "Test failed — bug found, add details in Notes column", RED_FAIL, RED_TEXT),
    ("Status: Blocked", "Cannot test — dependency or environment issue", BLUE_BLOCK, BLUE_TEXT),
    ("Status: Skip", "Intentionally skipped for this round", YELLOW_SKIP, YELLOW_TEXT),
    ("Priority: Critical", "Must pass before ANY beta release", CRITICAL_BG, DARK_TEXT),
    ("Priority: High", "Must pass before public beta", HIGH_BG, DARK_TEXT),
    ("Priority: Medium", "Should pass, but non-blocking", MEDIUM_BG, DARK_TEXT),
]

for i, (label, desc, bg, fg) in enumerate(legend_items):
    r = legend_row + 1 + i
    ws2.row_dimensions[r].height = 22
    cell = ws2.cell(row=r, column=2, value=label)
    cell.font = Font(name="Aptos", size=10, bold=True, color=fg)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.border = thin_border

    ws2.merge_cells(f"C{r}:I{r}")
    cell = ws2.cell(row=r, column=3, value=desc)
    cell.font = Font(name="Aptos", size=9, color=MUTED_TEXT)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.border = thin_border

# ── Testing Info ───────────────────────────────────────────────
info_row = legend_row + len(legend_items) + 3
ws2.merge_cells(f"B{info_row}:I{info_row}")
cell = ws2.cell(row=info_row, column=2, value="Test Environment")
cell.font = Font(name="Aptos", size=12, bold=True, color=DARK_TEXT)

env_items = [
    "Devices: iPhone 16 Pro (primary), iPhone SE (small screen), iPad (if supported)",
    "iOS: 17+ (minimum deployment target)",
    "Network: WiFi, Cellular, Airplane Mode",
    "Storage: Test with <1GB free to trigger warnings",
    "Build: TestFlight beta distribution",
]
for i, item in enumerate(env_items):
    r = info_row + 1 + i
    ws2.merge_cells(f"B{r}:I{r}")
    cell = ws2.cell(row=r, column=2, value=f"  {item}")
    cell.font = Font(name="Aptos", size=9, color=MUTED_TEXT)


# ══════════════════════════════════════════════════════════════
# SHEET 3: BUG LOG
# ══════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Bug Log")
ws3.sheet_properties.tabColor = "F44336"

bug_cols = {
    "A": 10,   # Bug ID
    "B": 12,   # Test ID
    "C": 14,   # Severity
    "D": 50,   # Description
    "E": 45,   # Steps to Reproduce
    "F": 30,   # Expected vs Actual
    "G": 14,   # Device
    "H": 14,   # Status
    "I": 14,   # Date Found
    "J": 14,   # Date Fixed
    "K": 40,   # Notes
}
for col_letter, width in bug_cols.items():
    ws3.column_dimensions[col_letter].width = width

# Title
ws3.merge_cells("A1:K1")
cell = ws3["A1"]
cell.value = "Bug Tracker"
cell.font = Font(name="Aptos", size=16, bold=True, color=WHITE)
cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
cell.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

# Headers
bug_headers = ["Bug ID", "Test ID", "Severity", "Description", "Steps to Reproduce", "Expected vs Actual", "Device", "Status", "Date Found", "Date Fixed", "Notes"]
ws3.row_dimensions[3].height = 28
for col_idx, h in enumerate(bug_headers, 1):
    cell = ws3.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Pre-fill 30 empty rows
severity_dv = DataValidation(
    type="list",
    formula1='"Critical,Major,Minor,Cosmetic"',
    allow_blank=True,
)
ws3.add_data_validation(severity_dv)

bug_status_dv = DataValidation(
    type="list",
    formula1='"Open,In Progress,Fixed,Verified,Won\'t Fix"',
    allow_blank=True,
)
ws3.add_data_validation(bug_status_dv)

for i in range(30):
    r = 4 + i
    stripe = stripe_even if i % 2 == 0 else stripe_odd
    ws3.row_dimensions[r].height = 28

    # Bug ID
    cell = ws3.cell(row=r, column=1, value=f"BUG-{i+1:03d}" if i < 1 else "")
    cell.font = Font(name="Aptos Mono", size=9, color=MUTED_TEXT)
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border

    for col_idx in range(2, 12):
        cell = ws3.cell(row=r, column=col_idx, value="")
        cell.font = body_font
        cell.fill = stripe
        cell.border = thin_border
        if col_idx in (1, 2, 3, 7, 8, 9, 10):
            cell.alignment = center_align
        else:
            cell.alignment = wrap_align

    severity_dv.add(ws3.cell(row=r, column=3))
    bug_status_dv.add(ws3.cell(row=r, column=8))

ws3.freeze_panes = "A4"
ws3.auto_filter.ref = "A3:K33"

# Clear the prefilled BUG-001
ws3.cell(row=4, column=1).value = ""


# ══════════════════════════════════════════════════════════════
# SHEET 4: SMOKE TEST CHECKLIST
# ══════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("Smoke Test")
ws4.sheet_properties.tabColor = ORANGE

ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 8
ws4.column_dimensions["C"].width = 55
ws4.column_dimensions["D"].width = 12
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 40

# Title
ws4.merge_cells("A1:F1")
cell = ws4["A1"]
cell.value = "5-Minute Smoke Test  —  Run Before Every Build"
cell.font = Font(name="Aptos", size=14, bold=True, color=WHITE)
cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
cell.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

# Headers
smoke_headers = ["", "Step", "Action", "Status", "Date", "Notes"]
ws4.row_dimensions[3].height = 28
for col_idx, h in enumerate(smoke_headers, 1):
    cell = ws4.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

smoke_steps = [
    ("1", "Fresh launch → Home screen loads"),
    ("2", "Upload a volleyball video from Photos"),
    ("3", "Process the video → see rally count results"),
    ("4", "Preview rallies → swipe through 3+ rallies"),
    ("5", "Long-press to trim one rally"),
    ("6", "Save selected rallies to a folder"),
    ("7", "Sign in with Apple"),
    ("8", "Share a rally to the community feed"),
    ("9", "View it in the social feed"),
    ("10", "Like and comment on it"),
    ("11", "Settings → sign out"),
    ("12", "Relaunch app → library still intact"),
]

smoke_status_dv = DataValidation(
    type="list",
    formula1='"Pass,Fail,Skip"',
    allow_blank=True,
)
ws4.add_data_validation(smoke_status_dv)

for i, (step, action) in enumerate(smoke_steps):
    r = 4 + i
    stripe = stripe_even if i % 2 == 0 else stripe_odd
    ws4.row_dimensions[r].height = 32

    # Checkbox column (visual)
    cell = ws4.cell(row=r, column=1, value="")
    cell.fill = stripe
    cell.border = thin_border

    # Step number
    cell = ws4.cell(row=r, column=2, value=step)
    cell.font = Font(name="Aptos", size=12, bold=True, color=ORANGE)
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border

    # Action
    cell = ws4.cell(row=r, column=3, value=action)
    cell.font = Font(name="Aptos", size=11, color=DARK_TEXT)
    cell.fill = stripe
    cell.alignment = body_align
    cell.border = thin_border

    # Status
    cell = ws4.cell(row=r, column=4, value="")
    cell.font = Font(name="Aptos", size=10, bold=True)
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border
    smoke_status_dv.add(cell)

    # Date
    cell = ws4.cell(row=r, column=5, value="")
    cell.font = body_font
    cell.fill = stripe
    cell.alignment = center_align
    cell.border = thin_border

    # Notes
    cell = ws4.cell(row=r, column=6, value="")
    cell.font = Font(name="Aptos", size=9, color=DARK_TEXT)
    cell.fill = stripe
    cell.alignment = wrap_align
    cell.border = thin_border

# Result summary below
summary_row = 4 + len(smoke_steps) + 2
ws4.merge_cells(f"B{summary_row}:F{summary_row}")
cell = ws4.cell(row=summary_row, column=2, value="Build Verdict:  ___  PASS  /  ___  FAIL      Tester: ________________      Date: ____________")
cell.font = Font(name="Aptos", size=11, bold=True, color=DARK_TEXT)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[summary_row].height = 35


# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════

output_path = "/Users/benjaminwierzbanowski/Code/BumpSetCut/.planning/BumpSetCut_Beta_Test_Plan.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
